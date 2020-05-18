import json
from openpyxl import load_workbook

for i in range(1, 5):
    XLSX_FILE = f"{i}.xlsx"
    JSON_FILE = f"../bank/squads/{i}/data.json"

    data = []
    workbook = load_workbook(XLSX_FILE)  # книга
    worksheet = workbook.active  # лист
    for row in worksheet.iter_rows(min_row=2):  # цикл по строкам (пропускаем шапку)
        data.append([r.value for r in row[1:]])  # добавляем список значений (пропускаем первую)
    base = {"data": data}
    json.dump(base, open(JSON_FILE, "w"))  # сериализуем список списков

